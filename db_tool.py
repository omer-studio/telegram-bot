#!/usr/bin/env python3
"""
🔍 MCP Database Tool - כלי לשאילתות read-only למסד הנתונים
מאפשר ל-Cursor Chat לבצע SELECT queries ישירות במסד הנתונים לצורך ניתוח

Usage from Cursor Chat:
@run_query("SELECT chat_id, code_try FROM user_profiles WHERE code_try>0 LIMIT 5")
"""

import os
import re
import json
import traceback
from typing import List, Dict, Any
import sys
from datetime import datetime
import csv
from io import StringIO

# ייבוא Excel libraries
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ Excel libraries not available - only CSV export supported")

try:
    import psycopg2
    import psycopg2.extras
    PSYCOPG2_AVAILABLE = True
except ImportError:
    print("⚠️ psycopg2 not available - database queries disabled")
    PSYCOPG2_AVAILABLE = False

# ייבוא logger בטוח
try:
    from simple_logger import logger
except ImportError:
    # ברירת מחדל אם logger לא זמין
    class SimpleLogger:
        def info(self, msg): print(f"ℹ️ {msg}")
        def error(self, msg): print(f"❌ {msg}")
        def warning(self, msg): print(f"⚠️ {msg}")
    logger = SimpleLogger()

def load_config():
    """🎯 טעינת קונפיגורציה דרך הפונקציה המרכזית"""
    try:
        from config import get_config
        return get_config()
    except Exception as e:
        print(f"❌ שגיאה בטעינת קונפיגורציה: {e}")
        return {}

def get_database_url():
    """קבלת URL למסד הנתונים מהקונפיגורציה הקיימת"""
    config = load_config()
    if not config:
        return None
    
    # שימוש באותו דפוס כמו בקוד הקיים
    return config.get("DATABASE_EXTERNAL_URL") or config.get("DATABASE_URL")

def run_query(query: str) -> List[Dict[str, Any]]:
    """
    Execute a read-only SQL SELECT on the production Postgres
    
    Args:
        query (str): SQL SELECT query to execute
        
    Returns:
        List[Dict[str, Any]]: Query results as list of dictionaries
        
    Raises:
        ValueError: If query is not a SELECT statement
        Exception: If database connection or query execution fails
    """
    if not PSYCOPG2_AVAILABLE:
        return [{"error": "psycopg2 not available - cannot execute database queries"}]
    
    # 🔒 אבטחה: רק שאילתות SELECT מותרות
    if not re.match(r'^\s*select', query.strip(), re.IGNORECASE):
        raise ValueError("❌ Only SELECT queries are allowed. Query must start with SELECT.")
    
    # חסימת פקודות מסוכנות - בדיקה על גבולות מילים
    dangerous_keywords = [
        'insert', 'update', 'delete', 'drop', 'create', 'alter', 
        'truncate', 'grant', 'revoke', 'exec', 'execute'
    ]
    
    query_lower = query.lower()
    for keyword in dangerous_keywords:
        # בדיקה שהמילה מופיעה כמילה שלמה ולא כחלק ממילה אחרת
        if re.search(r'\b' + keyword + r'\b', query_lower):
            raise ValueError(f"❌ Dangerous keyword '{keyword}' detected in query. Only SELECT queries allowed.")
    
    try:
        db_url = get_database_url()
        if not db_url:
            return [{"error": "❌ DATABASE_URL not found in configuration"}]
        
        print(f"🔍 Executing query: {query[:100]}{'...' if len(query) > 100 else ''}")
        
        # חיבור למסד הנתונים עם הגדרת SSL
        conn = psycopg2.connect(db_url, sslmode="require")
        
        # שימוש ב-RealDictCursor כדי לקבל תוצאות כ-dictionaries
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        # הרצת השאילתה
        cur.execute(query)
        
        # קבלת התוצאות
        rows = cur.fetchall()
        
        # סגירת החיבורים
        cur.close()
        conn.close()
        
        # המרה לרשימת dictionaries (JSON serializable)
        results = [dict(row) for row in rows]
        
        print(f"✅ Query executed successfully. Returned {len(results)} rows")
        
        return results
        
    except psycopg2.Error as db_error:
        error_msg = f"❌ Database error: {str(db_error)}"
        print(error_msg)
        return [{"error": error_msg, "type": "database_error"}]
        
    except Exception as e:
        error_msg = f"❌ Unexpected error: {str(e)}"
        print(error_msg)
        print(f"🔍 Traceback: {traceback.format_exc()}")
        return [{"error": error_msg, "type": "unexpected_error"}]

# פונקציות נוספות לשימוש נוח
def query_user_profiles(limit: int = 10) -> List[Dict[str, Any]]:
    """שאילתה מהירה לטבלת user_profiles"""
    query = f"SELECT * FROM user_profiles ORDER BY updated_at DESC LIMIT {limit}"
    return run_query(query)

def query_recent_messages(limit: int = 10) -> List[Dict[str, Any]]:
    """שאילתה מהירה להודעות אחרונות - רק שדות קיימים"""
    query = f"SELECT id, chat_id, user_msg, bot_msg, timestamp FROM chat_messages ORDER BY timestamp DESC LIMIT {limit}"
    return run_query(query)

def query_table_info(table_name: str) -> List[Dict[str, Any]]:
    """מידע על מבנה טבלה"""
    query = f"""
    SELECT column_name, data_type, is_nullable, column_default
    FROM information_schema.columns 
    WHERE table_name = '{table_name}'
    ORDER BY ordinal_position
    """
    return run_query(query)

# 🎯 פונקציות טבלאות מלאות בעברית
def טבלה_משתמשים(limit: int = 20) -> List[Dict[str, Any]]:
    """הצגת טבלת user_profiles מלאה עם כל השדות"""
    query = f"SELECT * FROM user_profiles ORDER BY updated_at DESC LIMIT {limit}"
    return run_query(query)

def טבלה_הודעות(limit: int = 30) -> List[Dict[str, Any]]:
    """הצגת טבלת chat_messages מלאה עם כל השדות הקיימים"""
    query = f"SELECT id, chat_id, user_msg, bot_msg, timestamp FROM chat_messages ORDER BY timestamp DESC LIMIT {limit}"
    return run_query(query)

def טבלה_הודעות_משתמש(chat_id: str, limit: int = 20) -> List[Dict[str, Any]]:
    """הצגת הודעות של משתמש ספציפי מלאות - רק שדות קיימים"""
    query = f"SELECT id, chat_id, user_msg, bot_msg, timestamp FROM chat_messages WHERE chat_id = '{chat_id}' ORDER BY timestamp DESC LIMIT {limit}"
    return run_query(query)

def טבלה_gpt_לוגים(limit: int = 25) -> List[Dict[str, Any]]:
    """🔥 הצגת טבלת interactions_log החדשה עם נתונים מתקדמים"""
    query = f"SELECT * FROM interactions_log ORDER BY timestamp DESC LIMIT {limit}"
    return run_query(query)

def טבלה_gpt_קריאות(limit: int = 25) -> List[Dict[str, Any]]:
    """🗑️ DEPRECATED: gpt_calls table disabled - השתמש ב-interactions_log במקום"""
    print("🔄 [DISABLED] gpt_calls table disabled - use interactions_log instead")
    return []

def טבלה_שגיאות(limit: int = 20) -> List[Dict[str, Any]]:
    """הצגת טבלת errors_stats מלאה עם כל השדות"""
    query = f"SELECT * FROM errors_stats ORDER BY timestamp DESC LIMIT {limit}"
    return run_query(query)

def טבלה_פריסות(limit: int = 15) -> List[Dict[str, Any]]:
    """הצגת טבלת deployment_logs מלאה עם כל השדות"""
    query = f"SELECT * FROM deployment_logs ORDER BY timestamp DESC LIMIT {limit}"
    return run_query(query)

def טבלה_משתמשים_קריטיים(limit: int = 15) -> List[Dict[str, Any]]:
    """הצגת טבלת critical_users מלאה עם כל השדות"""
    query = f"SELECT * FROM critical_users ORDER BY updated_at DESC LIMIT {limit}"
    return run_query(query)

def טבלה_חיובים(limit: int = 20) -> List[Dict[str, Any]]:
    """הצגת טבלת billing_usage מלאה עם כל השדות"""
    query = f"SELECT * FROM billing_usage ORDER BY timestamp DESC LIMIT {limit}"
    return run_query(query)

def טבלה_מגבלות_חינמיות(limit: int = 15) -> List[Dict[str, Any]]:
    """הצגת טבלת free_model_limits מלאה עם כל השדות"""
    query = f"SELECT * FROM free_model_limits ORDER BY updated_at DESC LIMIT {limit}"
    return run_query(query)

def טבלה_gpt_שימוש(limit: int = 20) -> List[Dict[str, Any]]:
    """הצגת טבלת gpt_usage_log מלאה עם כל השדות"""
    query = f"SELECT * FROM gpt_usage_log ORDER BY timestamp DESC LIMIT {limit}"
    return run_query(query)

def טבלה_מפורטת(table_name: str, limit: int = 20) -> List[Dict[str, Any]]:
    """הצגת כל טבלה לפי שם עם כל השדות"""
    query = f"SELECT * FROM {table_name} LIMIT {limit}"
    return run_query(query)

def חיפוש_משתמש_מלא(chat_id: str) -> Dict[str, Any]:
    """🔥 חיפוש מידע מלא על משתמש ספציפי - כל הטבלאות (עודכן ל-interactions_log)"""
    user_profile = run_query(f"SELECT * FROM user_profiles WHERE chat_id = '{chat_id}'")
    recent_messages = run_query(f"SELECT id, chat_id, user_msg, bot_msg, timestamp FROM chat_messages WHERE chat_id = '{chat_id}' ORDER BY timestamp DESC LIMIT 10")
    gpt_interactions = run_query(f"SELECT * FROM interactions_log WHERE chat_id = '{chat_id}' ORDER BY timestamp DESC LIMIT 10")
    
    return {
        "פרופיל_מלא": user_profile,
        "הודעות_מלאות": recent_messages,
        "אינטראקציות_מלאות": gpt_interactions
    }

def סטטיסטיקות_כלליות() -> Dict[str, Any]:
    """🔥 סטטיסטיקות כלליות של המערכת (עודכן ל-interactions_log)"""
    total_users = run_query("SELECT COUNT(*) as total FROM user_profiles")[0]['total']
    approved_users = run_query("SELECT COUNT(*) as approved FROM user_profiles WHERE approved = true")[0]['approved']
    messages_today = run_query("SELECT COUNT(*) as today FROM chat_messages WHERE DATE(timestamp) = CURRENT_DATE")[0]['today']
    # חישוב עלות מ-interactions_log (אגורות מומרות לדולרים)
    cost_week_agorot = run_query("SELECT COALESCE(SUM(total_cost_agorot), 0) as week_cost FROM interactions_log WHERE timestamp >= NOW() - INTERVAL '7 days'")[0]['week_cost']
    cost_week_usd = float(cost_week_agorot) / 100 / 3.7 if cost_week_agorot else 0  # 1 אגורה = 0.01 שקל, 1 שקל ≈ 0.27 דולר
    
    return {
        "סה_כ_משתמשים": total_users,
        "משתמשים_מאושרים": approved_users,
        "הודעות_היום": messages_today,
        "עלות_שבוע_דולר": round(cost_week_usd, 4)
    }

# פונקציה למתודולוגיות debug
def test_connection() -> Dict[str, Any]:
    """בדיקת חיבור למסד הנתונים"""
    try:
        result = run_query("SELECT NOW() as current_time, version() as pg_version")
        if result and not result[0].get("error"):
            return {
                "status": "✅ Connected successfully",
                "connection_test": result[0]
            }
        else:
            return {
                "status": "❌ Connection failed", 
                "error": result[0].get("error") if result else "Unknown error"
            }
    except Exception as e:
        return {
            "status": "❌ Connection test failed",
            "error": str(e)
        }

def export_table_to_utf8_csv(table_name, output_dir="csv_exports"):
    """
    מייצא טבלה לקובץ CSV עם UTF-8 BOM תקין 📊
    מבטיח תצוגה נכונה ב-Excel ובכל התוכנות
    """
    try:
        # יצירת תיקיה אם לא קיימת
        os.makedirs(output_dir, exist_ok=True)
        
        # קבלת URL למסד הנתונים
        db_url = get_database_url()
        if not db_url:
            logger.error("❌ לא נמצא URL למסד הנתונים")
            return None
            
        conn = psycopg2.connect(db_url, sslmode="require")
        cursor = conn.cursor()
        
        # שליפת כל הנתונים מהטבלה
        cursor.execute(f"SELECT * FROM {table_name}")
        rows = cursor.fetchall()
        
        # שליפת שמות העמודות
        if cursor.description:
            column_names = [desc[0] for desc in cursor.description]
        else:
            logger.error("❌ לא נמצאו עמודות בטבלה")
            return None
        
        # יצירת שם קובץ עם תאריך
        timestamp = datetime.now().strftime("%d_%m_%Y_%H%M")
        filename = f"{table_name}_{timestamp}.csv"
        filepath = os.path.join(output_dir, filename)
        
        # כתיבת הקובץ עם UTF-8 BOM
        with open(filepath, 'w', encoding='utf-8-sig', newline='') as csvfile:
            writer = csv.writer(csvfile)
            
            # כתיבת כותרות
            writer.writerow(column_names)
            
            # כתיבת הנתונים
            if rows:
                for row in rows:
                    # המרה של ערכים ל-string ותיקון None
                    processed_row = []
                    for value in row if row else []:
                        if value is None:
                            processed_row.append("")
                        elif isinstance(value, datetime):
                            processed_row.append(value.strftime("%Y-%m-%d %H:%M:%S"))
                        else:
                            processed_row.append(str(value))
                    writer.writerow(processed_row)
        
        cursor.close()
        conn.close()
        
        logger.info(f"✅ הטבלה {table_name} יוצאה בהצלחה: {filepath}")
        logger.info(f"📊 {len(rows)} שורות, {len(column_names)} עמודות")
        
        return filepath
        
    except Exception as e:
        logger.error(f"❌ שגיאה ביצוא הטבלה {table_name}: {e}")
        return None

def export_table_to_excel(table_name, output_dir="excel_exports"):
    """
    מייצא טבלה לקובץ Excel מעוצב ומקצועי 📊✨
    כולל עיצוב אוטומטי, פילטרים, וכותרות יפות
    """
    if not EXCEL_AVAILABLE:
        logger.error("❌ Excel libraries לא זמינות - השתמש ב-CSV במקום")
        return export_table_to_utf8_csv(table_name, output_dir.replace("excel_", "csv_"))
        
    try:
        # יצירת תיקיה אם לא קיימת
        os.makedirs(output_dir, exist_ok=True)
        
        # קבלת URL למסד הנתונים
        db_url = get_database_url()
        if not db_url:
            logger.error("❌ לא נמצא URL למסד הנתונים")
            return None
            
        conn = psycopg2.connect(db_url, sslmode="require")
        
        # שליפת הנתונים כ-DataFrame
        df = pd.read_sql_query(f"SELECT * FROM {table_name}", conn)
        conn.close()
        
        if df.empty:
            logger.warning(f"⚠️ הטבלה {table_name} ריקה")
            return None
        
        # יצירת שם קובץ עם תאריך
        timestamp = datetime.now().strftime("%d_%m_%Y_%H%M")
        filename = f"{table_name}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # יצירת Excel עם עיצוב מקצועי
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=table_name, index=False)
            
            # קבלת הגיליון לעיצוב
            worksheet = writer.sheets[table_name]
            
            # 🎨 עיצוב כותרות - כחול כהה מקצועי
            header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # 🎨 עיצוב גבולות
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            # החלת עיצוב על כותרות
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 🎨 עיצוב נתונים
            data_font = Font(name="Calibri", size=11)
            
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    cell.font = data_font
                    cell.alignment = Alignment(vertical="center", wrap_text=False)
                    
                    # צבע זברה לשורות - כחול בהיר
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
                    
                    # עיצוב מיוחד לתאריכים ומספרים
                    if isinstance(cell.value, datetime):
                        cell.number_format = 'DD/MM/YYYY HH:MM'
                    elif isinstance(cell.value, (int, float)) and cell.value != 0:
                        cell.number_format = '#,##0'
            
            # 📏 התאמת רוחב עמודות אוטומטית
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # הגבלת רוחב מקסימלי
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 🔒 הקפאת שורת כותרות
            worksheet.freeze_panes = "A2"
            
            # 🔽 הוספת פילטרים אוטומטיים
            max_col_letter = worksheet.cell(row=1, column=len(df.columns)).column_letter
            worksheet.auto_filter.ref = f"A1:{max_col_letter}{worksheet.max_row}"
            
            # 📊 הוספת טבלה מעוצבת
            table = Table(
                displayName=f"Table_{table_name}",
                ref=f"A1:{max_col_letter}{worksheet.max_row}"
            )
            
            # עיצוב טבלה
            style = TableStyleInfo(
                name="TableStyleMedium9", 
                showFirstColumn=False,
                showLastColumn=False, 
                showRowStripes=True, 
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            worksheet.add_table(table)
        
        logger.info(f"✅ הטבלה {table_name} יוצאה בהצלחה: {filepath}")
        logger.info(f"📊 {len(df)} שורות, {len(df.columns)} עמודות")
        logger.info(f"🎨 כולל עיצוב מקצועי: כותרות, פילטרים, זברה, והקפאת כותרות")
        
        return filepath
        
    except Exception as e:
        logger.error(f"❌ שגיאה ביצוא הטבלה {table_name} ל-Excel: {e}")
        return None

def export_all_main_tables():
    """מייצא את כל הטבלאות הראשיות ל-CSV עם UTF-8 תקין"""
    main_tables = ['user_profiles', 'chat_messages', 'interactions_log']
    exported_files = []
    
    logger.info("📤 מתחיל יצוא כל הטבלאות הראשיות...")
    
    for table in main_tables:
        filepath = export_table_to_utf8_csv(table)
        if filepath:
            exported_files.append(filepath)
    
    if exported_files:
        logger.info(f"✅ יוצאו {len(exported_files)} טבלאות בהצלחה:")
        for file in exported_files:
            logger.info(f"   📄 {file}")
    
    return exported_files

def export_all_main_tables_excel():
    """מייצא את כל הטבלאות הראשיות ל-Excel מעוצב ומקצועי"""
    main_tables = ['user_profiles', 'chat_messages', 'interactions_log']
    exported_files = []
    
    logger.info("📤 מתחיל יצוא כל הטבלאות הראשיות ל-Excel מעוצב...")
    
    for table in main_tables:
        filepath = export_table_to_excel(table)
        if filepath:
            exported_files.append(filepath)
    
    if exported_files:
        logger.info(f"✅ יוצאו {len(exported_files)} טבלאות Excel בהצלחה:")
        for file in exported_files:
            logger.info(f"   📄 {file}")
    
    return exported_files

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        command = sys.argv[1]
        
        if command == "export":
            # יצוא טבלה ספציפית או כל הטבלאות ל-CSV
            if len(sys.argv) > 2:
                table_name = sys.argv[2]
                print(f"📤 מייצא טבלה ל-CSV: {table_name}")
                filepath = export_table_to_utf8_csv(table_name)
                if filepath:
                    print(f"✅ הטבלה יוצאה בהצלחה: {filepath}")
                    print("🎯 הקובץ נשמר עם UTF-8 BOM - יפתח נכון ב-Excel!")
                else:
                    print("❌ שגיאה ביצוא הטבלה")
            else:
                print("📤 מייצא את כל הטבלאות הראשיות ל-CSV...")
                files = export_all_main_tables()
                if files:
                    print("🎯 כל הקבצים נשמרו עם UTF-8 BOM - יפתחו נכון ב-Excel!")
                    
        elif command == "excel":
            # יצוא טבלה ספציפית או כל הטבלאות ל-Excel מעוצב
            if len(sys.argv) > 2:
                table_name = sys.argv[2]
                print(f"📊 מייצא טבלה ל-Excel מעוצב: {table_name}")
                filepath = export_table_to_excel(table_name)
                if filepath:
                    print(f"✅ הטבלה יוצאה בהצלחה: {filepath}")
                    print("🎨 קובץ Excel מעוצב עם כותרות, פילטרים, זברה והקפאת כותרות!")
                else:
                    print("❌ שגיאה ביצוא הטבלה")
            else:
                print("📊 מייצא את כל הטבלאות הראשיות ל-Excel מעוצב...")
                files = export_all_main_tables_excel()
                if files:
                    print("🎨 כל הקבצים נשמרו כ-Excel מעוצב ומקצועי!")
                    
        elif command == "test":
            # בדיקת חיבור
            print("🧪 בודק חיבור למסד הנתונים...")
            connection_test = test_connection()
            print(f"תוצאה: {connection_test}")
            
        elif command == "stats":
            # סטטיסטיקות
            print("📊 מביא סטטיסטיקות...")
            stats = סטטיסטיקות_כלליות()
            print(f"סטטיסטיקות: {stats}")
            
        else:
            print("🔧 שימוש:")
            print("  python db_tool.py export [table_name]  # יצוא לCSV טבלה ספציפית או כל הטבלאות")
            print("  python db_tool.py excel [table_name]   # יצוא לExcel מעוצב טבלה ספציפית או כל הטבלאות")
            print("  python db_tool.py test                 # בדיקת חיבור")
            print("  python db_tool.py stats                # סטטיסטיקות כלליות")
    else:
        # בדיקה מקומית של הכלי
        print("🧪 Testing database tool...")
        
        # בדיקת חיבור
        connection_test = test_connection()
        print(f"Connection test: {connection_test}")
        
        # בדיקת שאילתה פשוטה
        try:
            result = run_query("SELECT COUNT(*) as total_users FROM user_profiles")
            print(f"Sample query result: {result}")
        except Exception as e:
            print(f"Sample query failed: {e}")
            
        print("\n🔧 לשימוש CLI:")
        print("  python db_tool.py export              # יצוא כל הטבלאות ל-CSV עם UTF-8 תקין")
        print("  python db_tool.py export user_profiles # יצוא טבלה ספציפית ל-CSV")
        print("  python db_tool.py excel               # יצוא כל הטבלאות ל-Excel מעוצב")
        print("  python db_tool.py excel user_profiles  # יצוא טבלה ספציפית ל-Excel מעוצב") 